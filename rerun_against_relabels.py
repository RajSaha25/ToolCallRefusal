#!/usr/bin/env python3
"""Re-run the label-dependent analyses against the three-way relabels in
relabel_out/relabel_<model>.csv.

Background. relabel_refusals.py rewrote the no-tool refusal labels with the shared
REFUSE/CAVEAT/COMPLY judge, so `new_refused` is True only for a clean refusal;
hedged compliance now lands in CAVEAT. It also recomputed tool-call safety with the
global-scope scorer (`tc_safe_fixed`). Both feed Table 1, but the mechanism results
-- projection AUC, the direction cosines, patching and steering -- were never rerun
on top of them. This script does that.

Stages:
  --stage cpu   re-scores cached generations only: label churn plus the Table 1
                columns under old / new / fixed labels. Reproduces
                relabel_out/summary_three_way.csv as a self-check. No GPU.
  --stage gpu   everything that needs residual-stream activations:
                  * r_text     harmful vs benign, no-tool        (request type)
                  * r_tool     harmful vs benign, tool-enabled   (request type, under tools)
                  * r_behav    refused vs complied harmful, no-tool, NEW labels
                                                                 (refusal decision)
                  * pairwise cosines among them -- the number main.tex:287 cites
                  * AUC of the projection against unsafe-action and against refusal,
                    under both old and new labels
                Needs the model in bf16 on a GPU.

Usage:
  python rerun_against_relabels.py --stage cpu --model Qwen3-14B
  python rerun_against_relabels.py --stage gpu --model Qwen3-14B --hf-id Qwen/Qwen3-14B
"""
import argparse
import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd

REPO = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO))
RESULTS = REPO / "results"
RELABEL = REPO / "relabel_out"
ART = REPO / "interp_artifacts"
OUT = REPO / "relabel_analysis"

# The eval dataset is gitignored; fall back to the copy in the main checkout.
_DATASET_NAMES = [
    "complete_combined_safety_eval_dataset_2304_normalized_labels.xlsx",
    "complete_safety_eval_dataset_normalized.xlsx",
]
# In a worktree the repo root is .claude/worktrees/<name>, so also look upward for
# the main checkout's (gitignored) data dir.
DATASET_CANDIDATES = [base / "data" / n
                      for base in [REPO, *REPO.parents[:4]]
                      for n in _DATASET_NAMES]

TRIPLE_KEY = ["batch", "domain", "scenario_id", "system_condition", "prompt_variant"]


# ------------------------------------------------------------------ data loading
def read_dataset():
    """ID -> (System Prompt, User Prompt). Read with openpyxl directly: the pinned
    pandas wants a newer openpyxl than the one installed here."""
    import openpyxl

    path = next((p for p in DATASET_CANDIDATES if p.exists()), None)
    if path is None:
        sys.exit("dataset xlsx not found; looked in:\n  " +
                 "\n  ".join(str(p) for p in DATASET_CANDIDATES))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(rows)]
    ix = {h: i for i, h in enumerate(hdr)}
    out = {}
    for r in rows:
        if r is None or all(v is None for v in r):
            continue
        out[str(r[ix["ID"]]).strip()] = (r[ix["System Prompt"]] or "",
                                         (r[ix["User Prompt"]] or "").replace("�", "—"))
    print(f"[data] {path.name}: {len(out)} prompts")
    return out


def tc_safe_fixed(row):
    """Re-score the saved tool calls with the global-scope scorer, matching
    relabel_refusals.tc_safe_fixed."""
    from tools.registry import get_domain

    if row["mode"] == "No-tool chat":
        return True
    try:
        calls = json.loads(row["tool_calls"]) if isinstance(row["tool_calls"], str) else []
    except json.JSONDecodeError:
        calls = []
    domain = get_domain(row["scenario_id"].split("_")[0])
    return len(domain.score_tool_calls_all(calls)) == 0


def load(model, with_prompts=False):
    """Full 2304-row results frame with the new labels merged onto the no-tool rows."""
    res = pd.read_csv(RESULTS / f"results_{model}.csv", low_memory=False)
    # `batch` is part of the triple key; relabel_refusals.add_batch derives it from the id.
    res["batch"] = (res["id"].astype(str).str.extract(r"(\d+)")[0].astype(int) - 1) // 576 + 1
    rel = pd.read_csv(RELABEL / f"relabel_{model}.csv", low_memory=False)
    lab = rel.set_index("id")[["new_label", "new_refused"]]
    res = res.join(lab, on="id")
    # Only no-tool rows were relabeled; tool rows keep their (unused) refusal label.
    nt = res["mode"] == "No-tool chat"
    res["refused_old"] = res["refused"].astype(bool)
    res["refused_new"] = res["new_refused"].where(nt).fillna(res["refused_old"]).astype(bool)
    res["tc_safe_fixed"] = res.apply(tc_safe_fixed, axis=1)
    missing = int(res.loc[nt, "new_label"].isna().sum())
    if missing:
        print(f"[warn] {missing} no-tool rows had no new label")
    if with_prompts:
        ds = read_dataset()
        res["system_prompt"] = res["id"].map(lambda i: ds.get(str(i).strip(), ("", ""))[0])
        res["user_prompt"] = res["id"].map(lambda i: ds.get(str(i).strip(), ("", ""))[1])
        if (res["user_prompt"] == "").any():
            print(f"[warn] {(res['user_prompt'] == '').sum()} rows got no prompt text")
    return res


# ---------------------------------------------------------------------- cpu stage
def divergence(df, refused_col, tc_col):
    """Conditioned divergence over matched harmful triples, identical to
    relabel_refusals.divergence."""
    h = df[df["request_type"] == "Harmful"]
    nt = h[h["mode"] == "No-tool chat"].groupby(TRIPLE_KEY)[refused_col].first().rename("ref")
    nn = (~h[h["mode"] == "Tool-enabled normal"].groupby(TRIPLE_KEY)[tc_col].first()).rename("un")
    ss = (~h[h["mode"] == "Tool-enabled stress"].groupby(TRIPLE_KEY)[tc_col].first()).rename("us")
    t = pd.concat([nt, nn, ss], axis=1).dropna()
    R = int(t["ref"].sum())
    if R == 0:
        return len(t), R, float("nan"), float("nan")
    return len(t), R, (t["ref"] & t["un"]).sum() / R, (t["ref"] & t["us"]).sum() / R


def stage_cpu(model):
    d = load(model)
    nt = d["mode"] == "No-tool chat"
    h = d[nt & (d["request_type"] == "Harmful")]
    b = d[nt & (d["request_type"] == "Benign")]

    flipped = h["refused_old"] != h["refused_new"]
    print(f"\n== {model} ==")
    print(f"harmful no-tool label churn: {int(flipped.sum())}/{len(h)} ({flipped.mean():.1%})")
    print(f"  three-way: {h['new_label'].value_counts().to_dict()}")

    _, R_old, dn_old, ds_old = divergence(d, "refused_old", "tc_safe")
    _, R_new, dn_new, ds_new = divergence(d, "refused_new", "tc_safe")
    _, _, dn_fix, ds_fix = divergence(d, "refused_new", "tc_safe_fixed")
    row = {
        "model": model,
        "ref_old": round(h["refused_old"].mean(), 3),
        "ref_new": round(h["refused_new"].mean(), 3),
        "overref_old": round(b["refused_old"].mean(), 3),
        "overref_new": round(b["refused_new"].mean(), 3),
        "div_n_old": round(dn_old, 3), "div_n_new": round(dn_new, 3), "div_n_fix": round(dn_fix, 3),
        "div_s_old": round(ds_old, 3), "div_s_new": round(ds_new, 3), "div_s_fix": round(ds_fix, 3),
        "R_old": R_old, "R_new": R_new,
    }
    print("  " + "  ".join(f"{k}={v}" for k, v in row.items() if k != "model"))

    # self-check against the committed summary
    ref = pd.read_csv(RELABEL / "summary_three_way.csv").set_index("model")
    if model in ref.index:
        r = ref.loc[model]
        bad = [k for k in ("ref_old", "ref_new", "div_n_old", "div_n_new",
                           "div_s_old", "div_s_new", "div_n_fix", "div_s_fix", "R_old", "R_new")
               if abs(float(row[k]) - float(r[k])) > 1e-6]
        print(f"  self-check vs summary_three_way.csv: "
              f"{'MATCH' if not bad else 'MISMATCH on ' + ', '.join(bad)}")
    OUT.mkdir(exist_ok=True)
    pd.DataFrame([row]).to_csv(OUT / f"table1_{model}.csv", index=False)
    return row


# ---------------------------------------------------------------------- gpu stage
def cfg_dims(cfg):
    """(hidden_size, num_hidden_layers), looking through a multimodal wrapper.

    Gemma-3 is a Gemma3Config whose text tower lives under .text_config, so the
    dimensions are not on the top-level config the way they are for Qwen/Mistral.
    """
    for c in (cfg, getattr(cfg, "text_config", None), getattr(cfg, "llm_config", None)):
        if c is not None and getattr(c, "hidden_size", None) and getattr(c, "num_hidden_layers", None):
            return int(c.hidden_size), int(c.num_hidden_layers)
    raise AttributeError(f"cannot find hidden_size/num_hidden_layers on {type(cfg).__name__}")


def auc_ci(pos, neg, n_boot=2000, seed=0):
    """AUC = P(score(pos) > score(neg)), Mann-Whitney form, bootstrap 95% CI."""
    pos, neg = np.asarray(pos, float), np.asarray(neg, float)
    if len(pos) < 2 or len(neg) < 2:
        return float("nan"), (float("nan"), float("nan")), len(pos), len(neg)

    def _auc(a, b):
        allv = np.concatenate([a, b])
        ranks = pd.Series(allv).rank().values
        return (ranks[:len(a)].sum() - len(a) * (len(a) + 1) / 2) / (len(a) * len(b))

    rng = np.random.RandomState(seed)
    boots = [_auc(rng.choice(pos, len(pos), True), rng.choice(neg, len(neg), True))
             for _ in range(n_boot)]
    return (float(_auc(pos, neg)),
            (round(float(np.percentile(boots, 2.5)), 3), round(float(np.percentile(boots, 97.5)), 3)),
            len(pos), len(neg))


def stage_gpu(model, hf_id, layer=None, n_dir=128, batch=8, out_name=None):
    import torch
    from transformers import AutoTokenizer, AutoModelForCausalLM
    from tools.registry import get_domain

    d = load(model, with_prompts=True)
    torch.set_grad_enabled(False)

    tok = AutoTokenizer.from_pretrained(hf_id, trust_remote_code=True)
    tok.padding_side = "left"
    if tok.pad_token is None:
        tok.pad_token = tok.eos_token
    net = AutoModelForCausalLM.from_pretrained(
        hf_id, torch_dtype=torch.bfloat16, device_map="cuda", trust_remote_code=True).eval()

    # Only Qwen's directions were committed (*.pt is gitignored), and a saved file
    # from another model has the wrong d_model, so treat it as optional: r_text is
    # rebuilt from activations either way, and the saved copy is only a cross-check.
    d_model, n_l = cfg_dims(net.config)
    saved = ART / model / "refusal_dirs.pt"
    if not saved.exists():
        saved = ART / "refusal_dirs.pt"
    r_text_saved = None
    if saved.exists():
        blob = torch.load(saved, map_location="cpu")
        if blob["dirs"].shape[-1] == d_model:
            if layer is None:
                lo, hi = int(0.35 * n_l), int(0.85 * n_l)
                layer = int(blob["sep"][lo:hi].argmax()) + lo
            r_text_saved = blob["dirs"][layer].float()
        else:
            print(f"[gpu] {saved.name} is for a different model "
                  f"(d_model {blob['dirs'].shape[-1]} != {d_model}); ignoring")
    if layer is None:
        sys.exit("no usable saved directions for this model -- pass --layer explicitly "
                 "(paper: Mistral 26/32, Command-R 26/32, Gemma 51/62, Llama 67/80)")
    print(f"[gpu] {model} layer={layer}/{n_l} d_model={d_model} "
          f"saved_dirs={'yes' if r_text_saved is not None else 'no'}")

    from render_utils import make_renderer
    render = make_renderer(tok, get_domain, log=print)
    res_native_tools = render.native_tools

    def resid(frame):
        """Last-token residual stream at `layer` (left padding, so index -1)."""
        prompts = [render(r) for _, r in frame.iterrows()]
        outs = []
        for i in range(0, len(prompts), batch):
            enc = tok(prompts[i:i + batch], return_tensors="pt", padding=True,
                      truncation=True, max_length=4096).to("cuda")
            hs = net(**enc, output_hidden_states=True).hidden_states[layer]
            outs.append(hs[:, -1, :].float().cpu())
        return torch.cat(outs)

    def unit(v):
        return v / v.norm()

    def cos(a, b):
        return round(float(torch.nn.functional.cosine_similarity(
            a.view(1, -1), b.view(1, -1)).item()), 3)

    nt = d["mode"] == "No-tool chat"
    tn = d["mode"] == "Tool-enabled normal"
    res = {"model": model, "layer": layer, "n_dir": n_dir,
           "tools_native": bool(res_native_tools)}

    # -- directions ------------------------------------------------------------
    hn = d[nt & (d["request_type"] == "Harmful")]
    bn = d[nt & (d["request_type"] == "Benign")]
    k = min(n_dir, len(hn), len(bn))
    A_h, A_b = resid(hn.sample(k, random_state=0)), resid(bn.sample(k, random_state=0))
    r_text = unit(A_h.mean(0) - A_b.mean(0))
    if r_text_saved is not None:
        res["cos_rtext_saved"] = cos(r_text, r_text_saved)
        print(f"[dir] r_text rebuilt (n={k}/class); cos vs saved = {res['cos_rtext_saved']}")
    else:
        print(f"[dir] r_text rebuilt (n={k}/class); no saved copy to cross-check")

    # Projection gap in this model's own activation units. run_scaled_evaluation
    # scales its addition coefficient (4x the gap) and steering grid to this, so
    # every model needs its own before ablation/addition/steering can be run.
    res["proj_gap"] = round(float((A_h @ r_text).mean() - (A_b @ r_text).mean()), 2)
    res["add_coef_4x"] = round(4 * abs(res["proj_gap"]), 1)
    print(f"[dir] proj_gap={res['proj_gap']}  -> addition coef (4x) = {res['add_coef_4x']}")

    ht, bt = d[tn & (d["request_type"] == "Harmful")], d[tn & (d["request_type"] == "Benign")]
    k2 = min(96, len(ht), len(bt))
    r_tool = unit(resid(ht.sample(k2, random_state=0)).mean(0)
                  - resid(bt.sample(k2, random_state=0)).mean(0))

    # behaviour-defined direction: refused vs complied, harmful no-tool.
    # `strict` drops CAVEAT entirely; `loose` folds CAVEAT in with complied.
    behav = {}
    for tag, comply_labels in (("strict", ["complied"]), ("loose", ["complied", "caveat"])):
        ref = hn[hn["new_label"] == "refused"]
        com = hn[hn["new_label"].isin(comply_labels)]
        k3 = min(n_dir, len(ref), len(com))
        print(f"[dir] r_behav[{tag}]: refused={len(ref)} complied={len(com)} -> n={k3}/class")
        if k3 < 8:
            print(f"[dir] r_behav[{tag}]: too few examples, skipped")
            continue
        behav[tag] = unit(resid(ref.sample(k3, random_state=0)).mean(0)
                          - resid(com.sample(k3, random_state=0)).mean(0))
        res[f"n_behav_{tag}"] = k3

    # old-label behaviour direction, for the before/after contrast
    ref_o, com_o = hn[hn["refused_old"]], hn[~hn["refused_old"]]
    k4 = min(n_dir, len(ref_o), len(com_o))
    r_behav_old = None
    if k4 >= 8:
        r_behav_old = unit(resid(ref_o.sample(k4, random_state=0)).mean(0)
                           - resid(com_o.sample(k4, random_state=0)).mean(0))
        res["n_behav_old"] = k4

    # -- cosines ---------------------------------------------------------------
    res["cos_rtext_rtool"] = cos(r_text, r_tool)
    for tag, v in behav.items():
        res[f"cos_rtext_rbehav_{tag}"] = cos(r_text, v)
        res[f"cos_rtool_rbehav_{tag}"] = cos(r_tool, v)
    if r_behav_old is not None:
        res["cos_rtext_rbehav_old"] = cos(r_text, r_behav_old)
        if "strict" in behav:
            res["cos_rbehav_old_new"] = cos(r_behav_old, behav["strict"])
    print("[cos] " + "  ".join(f"{k}={v}" for k, v in res.items() if k.startswith("cos_")))

    # -- AUC -------------------------------------------------------------------
    # (a) tool-mode harmful: does the projection predict the unsafe action?
    base = d[tn & (d["request_type"] == "Harmful")]
    base = base.sample(min(300, len(base)), random_state=0)
    proj = (resid(base) @ r_text).numpy()
    for tag, col in (("old", "tc_safe"), ("fixed", "tc_safe_fixed")):
        unsafe = ~base[col].astype(bool).values
        a, ci, np_, nn_ = auc_ci(proj[~unsafe], proj[unsafe])
        res[f"auc_unsafe_{tag}"] = {"auc": round(a, 3), "ci": ci, "n_safe": np_, "n_unsafe": nn_}
        print(f"[auc] proj vs unsafe [{tag:5s}]: {a:.3f} {ci}  (unsafe {nn_}/{len(base)})")

    # (b) no-tool harmful: does the projection predict the refusal decision?
    hb = d[nt & (d["request_type"] == "Harmful")]
    proj_nt = (resid(hb) @ r_text).numpy()
    for tag, col in (("old", "refused_old"), ("new", "refused_new")):
        lab = hb[col].astype(bool).values
        a, ci, np_, nn_ = auc_ci(proj_nt[lab], proj_nt[~lab])
        res[f"auc_refusal_{tag}"] = {"auc": round(a, 3), "ci": ci, "n_ref": np_, "n_comp": nn_}
        print(f"[auc] proj vs refusal [{tag:5s}]: {a:.3f} {ci}  (refused {np_}/{len(hb)})")

    OUT.mkdir(exist_ok=True)
    tosave = {"r_text": r_text, "r_tool": r_tool, "layer": layer}
    tosave.update({f"r_behav_{k}": v for k, v in behav.items()})
    if r_behav_old is not None:
        tosave["r_behav_old"] = r_behav_old
    torch.save(tosave, OUT / f"directions_{model}.pt")
    name = out_name or f"gpu_{model}.json"
    (OUT / name).write_text(json.dumps(res, indent=2))
    print(f"\nwrote {OUT / name} and directions_{model}.pt")
    return res


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--stage", choices=["cpu", "gpu"], required=True)
    ap.add_argument("--model", required=True)
    ap.add_argument("--hf-id")
    ap.add_argument("--layer", type=int)
    ap.add_argument("--n-dir", type=int, default=128)
    ap.add_argument("--batch", type=int, default=8)
    a = ap.parse_args()
    if a.stage == "cpu":
        stage_cpu(a.model)
    else:
        if not a.hf_id:
            sys.exit("--hf-id is required for the gpu stage")
        stage_gpu(a.model, a.hf_id, a.layer, a.n_dir, a.batch)
